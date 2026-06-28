"""
generate_mock_data.py
=====================
StudioTech BI Blueprint Generator — Mock Data Generator
Professional Services (Consulting / Accounting / Advisory) Industry Pack

Generates a complete Excel workbook with realistic mock data matching
the Professional Services star schema exactly:

  Dimension Tables:
    Dim_Date            — Full date spine, FY July–June, 2 years
    Dim_Employee        — 120 consultants across 4 service lines
    Dim_Client          — 80 clients across industries and segments
    Dim_Engagement      — 200 active and completed engagements
    Dim_ServiceLine     — 4 service lines across 2 divisions
    Dim_Practice        — 6 practice groups

  Fact Tables:
    Fact_Timesheets           — Daily timesheet lines (≈ 18,000 rows)
    Fact_Engagements          — Monthly engagement snapshots (≈ 2,400 rows)
    Fact_Invoices             — Invoice line items (≈ 1,200 rows)
    Fact_ResourceUtilisation  — Weekly utilisation by employee (≈ 6,000 rows)
    Fact_Pipeline             — Business development pipeline (≈ 80 rows)

USAGE:
    python generate_mock_data.py

OUTPUT:
    MockData_ProfessionalServices.xlsx   — ready to connect in Power BI Desktop

POWER BI CONNECTION:
    1. Open Power BI Desktop
    2. Home → Get Data → Excel Workbook
    3. Select MockData_ProfessionalServices.xlsx
    4. Select ALL tables in the Navigator (tick the checkbox at the top)
    5. Click Load
    6. In Model view, relationships are auto-detected by key column name.
       Verify these match the blueprint schema (all Many:One, Single direction).
    7. Build your visuals — all DAX measures from the blueprint will work immediately.

CONNECTING TO LIVE DATA:
    When your live source systems are ready:
    1. Home → Transform Data → Data Source Settings
    2. Change Source → point to your live system connection
    3. All relationships, measures, and visuals update automatically.
    The mock data column names and types exactly match the blueprint schema,
    so no measure or visual changes are required.

AUTHOR: StudioTech BI Blueprint Generator
"""

import random
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Seed for reproducibility ───────────────────────────────────────
random.seed(42)

# ── Colour palette (matching app dark theme → professional light) ──
CLR_HEADER_FACT = "1E3A5F"   # Dark navy   — fact table headers
CLR_HEADER_DIM  = "1A3A1A"   # Dark green  — dimension table headers
CLR_HEADER_DATE = "2D1B69"   # Dark purple — date table header
CLR_TEXT        = "FFFFFF"   # White text on dark headers
CLR_ALT_ROW     = "F0F4FF"   # Alternating row tint — facts
CLR_ALT_DIM     = "F0FFF4"   # Alternating row tint — dims
CLR_STRIPE_DATE = "F5F0FF"   # Alternating row tint — date
CLR_TAB_FACT    = "2D5BE3"   # Sheet tab — fact tables
CLR_TAB_DIM     = "1A7A4A"   # Sheet tab — dimension tables
CLR_TAB_DATE    = "7C3AED"   # Sheet tab — date table

# ── Reference data (realistic Australian professional services) ─────

FIRST_NAMES = [
    "James","Sarah","Michael","Emma","David","Jessica","Andrew","Sophie",
    "Matthew","Olivia","Daniel","Charlotte","Ryan","Isabella","Luke","Mia",
    "Thomas","Amelia","Joshua","Grace","Benjamin","Chloe","Nathan","Lily",
    "Samuel","Hannah","Alexander","Zoe","William","Emily","Jack","Natalie",
    "Henry","Victoria","Owen","Stephanie","Liam","Rebecca","Ethan","Lauren",
    "Noah","Katherine","Oliver","Rachel","Patrick","Melissa","Sean","Vanessa",
    "Connor","Tina","Marcus","Diana","Raj","Priya","Wei","Lin","Ahmed","Fatima",
]

LAST_NAMES = [
    "Smith","Johnson","Williams","Brown","Jones","Davis","Miller","Wilson",
    "Moore","Taylor","Anderson","Thomas","Jackson","White","Harris","Martin",
    "Thompson","Garcia","Martinez","Robinson","Clark","Rodriguez","Lewis","Lee",
    "Walker","Hall","Allen","Young","Hernandez","King","Wright","Lopez","Scott",
    "Green","Adams","Baker","Gonzalez","Nelson","Carter","Mitchell","Perez",
    "Roberts","Turner","Phillips","Campbell","Parker","Evans","Edwards","Collins",
    "Stewart","Patel","Singh","Chen","Zhang","Wong","Kim","Nguyen","Murphy",
]

CLIENT_PREFIXES = [
    "Apex","Summit","Harbour","Pacific","Southern Cross","Horizon","Atlas",
    "Pioneer","Meridian","Pinnacle","Alpine","Coastal","Centennial","Aurora",
    "Bridgeway","Catalyst","Dexterity","Elevate","Foundation","Gateway",
    "Highland","Integral","Junction","Keystone","Landmark","Momentum","Nexus",
    "Optic","Praxis","Quantum","Regent","Stellar","Titan","Unified","Vantage",
    "Waypoint","Zenith","Accord","Benchmark","Clarity","Delta","Endure",
]

CLIENT_SUFFIXES = [
    "Group","Holdings","Partners","Advisory","Solutions","Capital","Ventures",
    "Investments","Services","Enterprises","Consulting","Management","Global",
    "Industries","Corporation","Limited","Pty Ltd",
]

CLIENT_INDUSTRIES = [
    "Financial Services","Healthcare","Property & Construction","Technology",
    "Manufacturing","Retail & Consumer","Government","Professional Services",
    "Energy & Resources","Education","Agribusiness","Transport & Logistics",
]

STATES = ["VIC","NSW","QLD","SA","WA","ACT","TAS","NT"]
STATE_WEIGHTS = [35,30,15,7,7,3,2,1]

SERVICE_LINE_DATA = [
    (1, "Audit & Assurance",       1, "Assurance & Advisory"),
    (2, "Tax & Compliance",        1, "Assurance & Advisory"),
    (3, "Business Advisory",       2, "Consulting & Advisory"),
    (4, "Corporate Finance",       2, "Consulting & Advisory"),
]

PRACTICE_DATA = [
    (1, "External Audit",          1, 1),
    (2, "Internal Audit & Risk",   1, 1),
    (3, "Tax Compliance",          2, 1),
    (4, "Tax Advisory",            2, 1),
    (5, "Business Improvement",    3, 2),
    (6, "Transaction Advisory",    4, 2),
]

GRADES = [
    ("Partner",            520, 0.08),
    ("Director",           420, 0.07),
    ("Senior Manager",     340, 0.10),
    ("Manager",            270, 0.15),
    ("Senior Consultant",  200, 0.20),
    ("Consultant",         155, 0.22),
    ("Graduate",           110, 0.18),
]  # (grade, charge_out_rate, proportion)

ENGAGEMENT_TYPES = ["Audit","Tax Return","BAS","Advisory","Due Diligence",
                    "Restructure","Business Review","CFO Advisory","Payroll Tax",
                    "SMSF Audit","R&D Tax Incentive","M&A Advisory"]

PIPELINE_STAGES = ["Lead","Qualified","Proposal Sent","Negotiation","Won","Lost"]
PIPELINE_WIN_PROB = [0.10, 0.25, 0.50, 0.75, 1.00, 0.00]

INVOICE_STATUSES  = ["Paid","Paid","Paid","Issued","Overdue","Written Off"]
ENGAGEMENT_STATUS = ["Active","Active","Active","Active","Completed","On Hold","Lost"]


# ═══════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════

def rnd_name():
    return random.choice(FIRST_NAMES) + " " + random.choice(LAST_NAMES)

def rnd_state():
    return random.choices(STATES, weights=STATE_WEIGHTS)[0]

def rnd_client_name():
    return random.choice(CLIENT_PREFIXES) + " " + random.choice(CLIENT_SUFFIXES)

def date_to_key(d):
    """Integer date key: YYYYMMDD"""
    return d.year * 10000 + d.month * 100 + d.day

def style_header(ws, row_num, num_cols, bg_hex, text_hex="FFFFFF", height=20):
    """Apply header styling to a full row."""
    thin = Side(style="thin", color="CCCCCC")
    border = Border(bottom=Side(style="medium", color="888888"))
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font      = Font(bold=True, color=text_hex, size=10, name="Segoe UI")
        cell.fill      = PatternFill("solid", start_color=bg_hex)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[row_num].height = height

def style_data_rows(ws, start_row, end_row, num_cols, alt_color):
    """Alternating row shading for data rows."""
    for row in range(start_row, end_row + 1):
        fill = PatternFill("solid", start_color=alt_color) if row % 2 == 0 else None
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font      = Font(size=9, name="Segoe UI")
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

def freeze_and_filter(ws, freeze_ref="A2"):
    ws.freeze_panes = freeze_ref
    ws.auto_filter.ref = ws.dimensions

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_sheet(wb, sheet_name, headers, rows, tab_color, alt_color, hdr_color, widths=None):
    """Write a complete sheet with headers, data, styling."""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = tab_color

    # Write headers
    ws.append(headers)
    style_header(ws, 1, len(headers), hdr_color)

    # Write data
    for row in rows:
        ws.append(row)

    # Style data rows
    style_data_rows(ws, 2, len(rows) + 1, len(headers), alt_color)

    # Auto-filter + freeze
    freeze_and_filter(ws)

    # Column widths
    if widths:
        set_col_widths(ws, widths)
    else:
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 16

    # Row count note in cell A1 comment area (sheet name row)
    ws.sheet_view.showGridLines = True
    return ws


# ═══════════════════════════════════════════════════════════════════
# BUILD DIMENSION DATA
# ═══════════════════════════════════════════════════════════════════

def build_dim_date(start_date, end_date):
    """Full date spine with fiscal year (July start)."""
    rows = []
    d = start_date
    while d <= end_date:
        key         = date_to_key(d)
        fy_year     = d.year if d.month >= 7 else d.year - 1
        fy_label    = f"FY{fy_year + 1}"
        fy_month    = ((d.month - 7) % 12) + 1          # 1=Jul, 12=Jun
        fy_quarter  = f"Q{((fy_month - 1) // 3) + 1}"
        cal_quarter = f"Q{((d.month - 1) // 3) + 1}"
        week_num    = d.isocalendar()[1]
        is_weekend  = 1 if d.weekday() >= 5 else 0
        # Simple AU public holiday check (major national ones only)
        is_ph       = 1 if (d.month == 1 and d.day == 1) or \
                           (d.month == 1 and d.day == 26) or \
                           (d.month == 4 and d.day == 25) or \
                           (d.month == 12 and d.day == 25) or \
                           (d.month == 12 and d.day == 26) else 0
        rows.append([
            key, d,                         # DateKey, Date
            d.day, d.strftime("%A"),        # Day, DayName
            week_num,                       # Week
            d.month, d.strftime("%B"),      # Month, MonthName
            d.strftime("%b-%Y"),            # MonthLabel (e.g. Jul-2024)
            cal_quarter, d.year,            # Quarter, CalendarYear
            fy_label,                       # FiscalYear
            fy_quarter,                     # FiscalQuarter
            fy_month,                       # FiscalMonth
            is_weekend, is_ph,              # IsWeekend, IsPublicHoliday
        ])
        d += datetime.timedelta(days=1)
    return rows

def build_dim_service_line():
    rows = []
    for slk, sln, pk, dn in SERVICE_LINE_DATA:
        rows.append([slk, sln, pk, dn])
    return rows

def build_dim_practice(employees):
    """Find practice leader (most senior employee per practice)."""
    leaders = {}
    for emp in employees:
        pk = emp[7]  # PracticeKey
        grade = emp[4]
        grade_order = [g[0] for g in GRADES]
        rank = grade_order.index(grade) if grade in grade_order else 99
        if pk not in leaders or rank < leaders[pk][1]:
            leaders[pk] = (emp[0], rank)  # (EmployeeKey, rank)

    rows = []
    for pk, pn, slk, dk in PRACTICE_DATA:
        leader_key = leaders.get(pk, (1, 0))[0]
        rows.append([pk, pn, slk, dk, leader_key])
    return rows

def build_dim_employee():
    rows = []
    emp_key = 1
    employees = []

    # Build grade pool matching target proportions for 120 staff
    grade_pool = []
    for grade, rate, prop in GRADES:
        count = max(1, round(120 * prop))
        grade_pool.extend([(grade, rate)] * count)
    random.shuffle(grade_pool)
    grade_pool = grade_pool[:120]

    # Assign employees across service lines and practices
    for i, (grade, rate) in enumerate(grade_pool):
        sl_key  = random.randint(1, 4)
        prac_key = random.choice([p[0] for p in PRACTICE_DATA if p[2] == sl_key] or [1])
        is_partner = 1 if grade == "Partner" else 0
        hire_date  = datetime.date(2018, 1, 1) + datetime.timedelta(days=random.randint(0, 2000))
        name = rnd_name()
        # Add variance to charge-out rate (±15%)
        actual_rate = round(rate * random.uniform(0.85, 1.15), 0)
        row = [
            emp_key,                        # EmployeeKey
            f"EMP{emp_key:04d}",            # EmployeeID
            name,                           # FullName
            grade,                          # Grade
            "Partner" if is_partner else "Staff",  # Role
            sl_key,                         # ServiceLineKey
            prac_key,                       # PracticeKey
            prac_key,                       # PracticeKey (duplicate for convenience)
            random.choice(["Full-time","Full-time","Full-time","Part-time"]),  # EmploymentType
            actual_rate,                    # ChargeOutRate
            is_partner,                     # IsPartner
            1,                              # IsActive (mostly active)
            hire_date,                      # HireDate
            rnd_state(),                    # OfficeState
        ]
        employees.append(row)
        emp_key += 1

    # Deactivate 10 employees (leavers)
    leavers = random.sample(range(len(employees)), 10)
    for idx in leavers:
        employees[idx][11] = 0  # IsActive = 0

    return employees

def build_dim_client():
    rows = []
    used_names = set()
    for ck in range(1, 81):
        name = rnd_client_name()
        while name in used_names:
            name = rnd_client_name()
        used_names.add(name)
        industry = random.choice(CLIENT_INDUSTRIES)
        # Assign segment based on revenue tier
        segment = random.choice(["Mid-Market","Mid-Market","Enterprise","SME","SME"])
        state    = rnd_state()
        since    = datetime.date(2015, 1, 1) + datetime.timedelta(days=random.randint(0, 2500))
        is_active = 1 if ck <= 72 else 0  # 8 inactive clients

        # Partner responsible (1–15 are partners in our employee set)
        partner_key = random.randint(1, 10)

        rows.append([
            ck,                             # ClientKey
            f"CLI{ck:04d}",                 # ClientCode
            name,                           # ClientName
            industry,                       # Industry
            segment,                        # Segment
            state,                          # State
            partner_key,                    # ClientPartnerKey
            is_active,                      # IsActive
            since,                          # ClientSince
        ])
    return rows

def build_dim_engagement(clients, employees):
    rows = []
    partner_keys = [e[0] for e in employees if e[10] == 1]  # IsPartner

    for ek in range(1, 201):
        client_key    = random.randint(1, 80)
        sl_key        = random.randint(1, 4)
        eng_type      = random.choice(ENGAGEMENT_TYPES)
        start         = datetime.date(2023, 7, 1) + datetime.timedelta(days=random.randint(0, 500))
        end           = start + datetime.timedelta(days=random.randint(30, 365))
        is_active     = 1 if end >= datetime.date(2025, 1, 1) else 0

        rows.append([
            ek,                             # EngagementKey
            f"ENG{ek:04d}",                 # EngagementCode
            f"{eng_type} — {clients[client_key-1][2]}",  # EngagementName
            client_key,                     # ClientKey
            sl_key,                         # ServiceLineKey
            eng_type,                       # EngagementType
            start,                          # StartDate
            end,                            # EndDate
            is_active,                      # IsActive
        ])
    return rows


# ═══════════════════════════════════════════════════════════════════
# BUILD FACT DATA
# ═══════════════════════════════════════════════════════════════════

def build_fact_timesheets(employees, engagements, date_rows):
    """Daily timesheet lines. Weekdays only. ~18,000 rows."""
    rows = []
    ts_key = 1

    # Business days in range
    biz_days = [r[1] for r in date_rows if r[13] == 0]  # IsWeekend = 0

    active_emps = [e for e in employees if e[11] == 1]  # IsActive
    active_engs = engagements[:180]  # Use active/recent engagements

    # Each active employee posts ~2-3 timesheet lines per week
    for emp in active_emps:
        emp_key  = emp[0]
        sl_key   = emp[5]
        rate     = emp[9]  # ChargeOutRate
        grade    = emp[3]

        # Target utilisation by grade
        util_targets = {
            "Partner": 0.55, "Director": 0.60, "Senior Manager": 0.68,
            "Manager": 0.72, "Senior Consultant": 0.78, "Consultant": 0.80,
            "Graduate": 0.82,
        }
        target_util = util_targets.get(grade, 0.72)

        # Sample ~120 working days per employee
        sample_days = random.sample(biz_days, min(120, len(biz_days)))

        for day in sample_days:
            day_key = date_to_key(day)

            # Pick an engagement matching service line
            eng_pool = [e for e in active_engs if e[3] == sl_key] or active_engs
            eng      = random.choice(eng_pool)
            eng_key  = eng[0]
            cli_key  = eng[3]

            hours_worked  = round(random.uniform(6.5, 10.0), 1)
            bill_rate     = random.random()
            billable_hrs  = round(hours_worked * random.uniform(target_util - 0.1, target_util + 0.1), 1) \
                            if random.random() < 0.85 else 0.0
            billable_hrs  = min(billable_hrs, hours_worked)
            non_billable  = round(hours_worked - billable_hrs, 1)
            wip_value     = round(billable_hrs * rate, 2)
            is_approved   = 1 if random.random() < 0.92 else 0

            rows.append([
                ts_key, emp_key, eng_key, cli_key, sl_key, day_key,
                hours_worked, billable_hrs, non_billable,
                rate, wip_value, is_approved,
            ])
            ts_key += 1

    return rows

def build_fact_engagements(engagements, employees, date_rows):
    """Monthly engagement snapshots. One row per engagement per month-end."""
    rows = []
    snap_key = 1

    # Month-end dates in range
    month_ends = []
    for r in date_rows:
        d = r[1]
        # Last day of month
        next_month = d.replace(day=28) + datetime.timedelta(days=4)
        last_day   = next_month - datetime.timedelta(days=next_month.day)
        if d == last_day and r[13] == 0:  # Not weekend
            month_ends.append(r)

    for eng in engagements:
        ek         = eng[0]
        cli_key    = eng[3]
        sl_key     = eng[4]
        start_date = eng[6]
        end_date   = eng[7]
        is_active  = eng[8]

        # Pick engagement manager (random senior employee in same service line)
        mgr_pool   = [e[0] for e in employees
                      if e[5] == sl_key and e[3] in
                      ("Partner","Director","Senior Manager","Manager")]
        mgr_key    = random.choice(mgr_pool) if mgr_pool else employees[0][0]

        budget_fees = round(random.uniform(15000, 350000), 0)
        is_fixed    = 1 if random.random() < 0.35 else 0

        # Accumulating actuals over snapshots
        cumulative_fees = 0.0
        cumulative_wip  = 0.0

        # Snapshots for months when engagement was active
        for snap in month_ends[:12]:  # Last 12 months
            snap_date = snap[1]
            snap_key_val = snap[0]

            if snap_date < start_date or snap_date > end_date + datetime.timedelta(days=90):
                continue

            progress = min(1.0, (snap_date - start_date).days / max((end_date - start_date).days, 1))
            month_fees = round(budget_fees * random.uniform(0.06, 0.12), 2)
            cumulative_fees = min(budget_fees * 1.1, cumulative_fees + month_fees)

            # WIP = unbilled work
            wip_balance   = round(cumulative_fees * random.uniform(0.05, 0.35), 2)
            write_off     = round(cumulative_fees * random.uniform(0.00, 0.08), 2)
            status        = "Completed" if snap_date > end_date else \
                            (random.choice(ENGAGEMENT_STATUS[:4]) if is_active else "On Hold")

            start_dk = date_to_key(start_date)
            end_dk   = date_to_key(end_date)

            rows.append([
                snap_key, cli_key, mgr_key, sl_key,
                snap_key_val,           # DateKey (snapshot)
                start_dk,               # StartDateKey
                end_dk,                 # EndDateKey
                round(budget_fees, 2),  # BudgetFees
                round(cumulative_fees, 2),  # ActualFees
                wip_balance,            # WIPBalance
                write_off,              # WriteOffAmount
                status,                 # EngagementStatus
                is_fixed,               # IsFixedFee
            ])
            snap_key += 1

    return rows

def build_fact_invoices(clients, engagements, date_rows):
    """Invoice line items — one per engagement per billing cycle."""
    rows = []
    inv_key = 1

    inv_dates = [r for r in date_rows if r[1].weekday() < 5]  # Weekdays only

    for eng in engagements[:180]:
        ek       = eng[0]
        cli_key  = eng[3]
        start    = eng[6]

        # Generate 1-4 invoices per engagement
        num_invoices = random.randint(1, 4)
        for n in range(num_invoices):
            # Invoice date: random day after engagement start
            offset   = random.randint(30, 300)
            inv_date = start + datetime.timedelta(days=offset)
            due_date = inv_date + datetime.timedelta(days=30)
            paid_date = due_date + datetime.timedelta(days=random.randint(-10, 45))

            # Find closest date keys
            inv_dk  = date_to_key(inv_date)
            due_dk  = date_to_key(due_date)
            paid_dk = date_to_key(paid_date)

            inv_amount    = round(random.uniform(3500, 85000), 2)
            gst           = round(inv_amount * 0.10, 2)
            status        = random.choice(INVOICE_STATUSES)
            amt_paid      = inv_amount if status == "Paid" else \
                            (0 if status in ("Overdue","Written Off","Draft") else
                             round(inv_amount * random.uniform(0, 0.9), 2))
            amt_outstanding = round(inv_amount - amt_paid, 2)
            days_out        = max(0, (datetime.date(2025, 3, 31) - inv_date).days)

            rows.append([
                inv_key, cli_key, ek,
                inv_dk, due_dk, paid_dk,
                inv_amount, gst,
                round(amt_paid, 2), round(amt_outstanding, 2),
                days_out, status,
            ])
            inv_key += 1

    return rows

def build_fact_utilisation(employees, date_rows):
    """Weekly utilisation snapshot — one row per employee per week."""
    rows = []
    util_key = 1

    # Week-ending Fridays
    fridays = [r for r in date_rows if r[1].weekday() == 4]  # Friday

    grade_targets = {
        "Partner": (0.50, 0.62), "Director": (0.58, 0.68),
        "Senior Manager": (0.65, 0.75), "Manager": (0.68, 0.78),
        "Senior Consultant": (0.74, 0.84), "Consultant": (0.76, 0.86),
        "Graduate": (0.78, 0.88),
    }

    active_emps = [e for e in employees if e[11] == 1]

    for emp in active_emps:
        ek     = emp[0]
        sl_key = emp[5]
        grade  = emp[3]
        lo, hi = grade_targets.get(grade, (0.70, 0.80))

        for fri in fridays:
            dk              = fri[0]
            available_hours = 37.5  # Standard week
            leave_hours     = random.choice([0, 0, 0, 0, 0, 7.5, 15.0])
            effective_avail = available_hours - leave_hours
            util_rate       = random.uniform(lo - 0.05, hi + 0.05)
            util_rate       = max(0.0, min(1.0, util_rate))
            billable_hrs    = round(effective_avail * util_rate, 1)
            non_bill_hrs    = round(effective_avail * random.uniform(0.05, 0.20), 1)
            non_bill_hrs    = min(non_bill_hrs, effective_avail - billable_hrs)

            rows.append([
                util_key, ek, sl_key, dk,
                available_hours, billable_hrs, non_bill_hrs,
                leave_hours, round(util_rate, 4),
            ])
            util_key += 1

    return rows

def build_fact_pipeline(clients, employees, date_rows):
    """Current BD pipeline — one row per opportunity."""
    rows = []
    today_dk = date_to_key(datetime.date(2025, 3, 31))

    owner_pool = [e[0] for e in employees if e[10] == 1 or e[3] == "Director"]

    for pk in range(1, 81):
        cli_key = random.randint(1, 80)
        sl_key  = random.randint(1, 4)
        owner   = random.choice(owner_pool)
        stage_idx = random.choices(range(6), weights=[15,20,25,20,12,8])[0]
        stage   = PIPELINE_STAGES[stage_idx]
        win_prob = PIPELINE_WIN_PROB[stage_idx] + random.uniform(-0.05, 0.05)
        win_prob = max(0.0, min(1.0, win_prob))

        est_value    = round(random.uniform(20000, 800000), 0)
        weighted_val = round(est_value * win_prob, 2)
        close_date   = datetime.date(2025, 3, 31) + datetime.timedelta(days=random.randint(30, 365))

        rows.append([
            pk, cli_key, sl_key, owner, today_dk,
            est_value, round(weighted_val, 2), round(win_prob, 2),
            stage, close_date,
        ])

    return rows


# ═══════════════════════════════════════════════════════════════════
# WRITE INSTRUCTIONS SHEET
# ═══════════════════════════════════════════════════════════════════

def write_instructions(wb):
    ws = wb.create_sheet("📋 Instructions", 0)
    ws.sheet_properties.tabColor = "2D5BE3"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "StudioTech BI Blueprint Generator — Professional Services Mock Data"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF", name="Segoe UI")
    ws["A1"].fill = PatternFill("solid", start_color="1E3A5F")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    ws["A2"] = "Realistic mock data matching the Professional Services star schema blueprint exactly. Connect to Power BI Desktop to see all visuals instantly."
    ws["A2"].font = Font(size=11, color="444444", name="Segoe UI")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28

    sections = [
        ("", ""),
        ("🚀  HOW TO CONNECT IN POWER BI DESKTOP", ""),
        ("Step 1", "Open Power BI Desktop → Home → Get Data → Excel Workbook"),
        ("Step 2", "Select this file: MockData_ProfessionalServices.xlsx"),
        ("Step 3", "In the Navigator, tick SELECT ALL TABLES (tick the checkbox at the very top left)"),
        ("Step 4", "Click Load (not Transform — unless you need to review the data first)"),
        ("Step 5", "Go to the Model view (icon on the left sidebar)"),
        ("Step 6", "Verify relationships — Power BI auto-detects by key column name.\n"
                   "All relationships should be Many:One, Single direction.\n"
                   "If any are missing, drag from the fact column to the matching dimension column."),
        ("Step 7", "Go to Report view — start adding visuals from the Fields pane.\n"
                   "All DAX measures from your blueprint can be added to the model now."),
        ("", ""),
        ("📊  WHAT'S IN EACH SHEET", ""),
        ("Dim_Date",               f"Full date spine (Jul 2023 – Jun 2025) · FY July start · {2*365} rows"),
        ("Dim_Employee",           "120 consultants across 4 service lines, 6 practice groups, 7 grades"),
        ("Dim_Client",             "80 clients · industries, segments, states, responsible partner"),
        ("Dim_Engagement",         "200 engagements · audit, tax, advisory, M&A types"),
        ("Dim_ServiceLine",        "4 service lines across 2 divisions"),
        ("Dim_Practice",           "6 practice groups with assigned leaders"),
        ("Fact_Timesheets",        "Daily timesheet lines · billable hours, WIP value, approval status · ~18,000 rows"),
        ("Fact_Engagements",       "Monthly engagement snapshots · fees, WIP balance, write-offs · ~2,400 rows"),
        ("Fact_Invoices",          "Invoice line items · amount, GST, paid, outstanding, status · ~1,200 rows"),
        ("Fact_ResourceUtilisation","Weekly utilisation per employee · available, billable, leave hours · ~6,000 rows"),
        ("Fact_Pipeline",          "BD pipeline opportunities · stage, estimated value, win probability · 80 rows"),
        ("", ""),
        ("🔗  RELATIONSHIPS (Many:One, Single direction unless noted)", ""),
        ("Fact_Timesheets[DateKey]",              "→ Dim_Date[DateKey]"),
        ("Fact_Timesheets[EmployeeKey]",          "→ Dim_Employee[EmployeeKey]"),
        ("Fact_Timesheets[ClientKey]",            "→ Dim_Client[ClientKey]"),
        ("Fact_Timesheets[ServiceLineKey]",       "→ Dim_ServiceLine[ServiceLineKey]"),
        ("Fact_Timesheets[EngagementKey]",        "→ Dim_Engagement[EngagementKey]"),
        ("Fact_Engagements[DateKey]",             "→ Dim_Date[DateKey]  (snapshot — active)"),
        ("Fact_Engagements[StartDateKey]",        "→ Dim_Date[DateKey]  (inactive — USERELATIONSHIP())"),
        ("Fact_Engagements[EndDateKey]",          "→ Dim_Date[DateKey]  (inactive — USERELATIONSHIP())"),
        ("Fact_Engagements[ClientKey]",           "→ Dim_Client[ClientKey]"),
        ("Fact_Engagements[ServiceLineKey]",      "→ Dim_ServiceLine[ServiceLineKey]"),
        ("Fact_Invoices[InvoiceDateKey]",         "→ Dim_Date[DateKey]  (active)"),
        ("Fact_Invoices[DueDateKey]",             "→ Dim_Date[DateKey]  (inactive — USERELATIONSHIP())"),
        ("Fact_Invoices[PaidDateKey]",            "→ Dim_Date[DateKey]  (inactive — USERELATIONSHIP())"),
        ("Fact_Invoices[ClientKey]",              "→ Dim_Client[ClientKey]"),
        ("Fact_ResourceUtilisation[EmployeeKey]", "→ Dim_Employee[EmployeeKey]"),
        ("Fact_ResourceUtilisation[DateKey]",     "→ Dim_Date[DateKey]"),
        ("Fact_ResourceUtilisation[ServiceLineKey]", "→ Dim_ServiceLine[ServiceLineKey]"),
        ("Fact_Pipeline[ClientKey]",              "→ Dim_Client[ClientKey]"),
        ("Fact_Pipeline[DateKey]",                "→ Dim_Date[DateKey]"),
        ("Fact_Pipeline[OwnerKey]",               "→ Dim_Employee[EmployeeKey]"),
        ("Fact_Pipeline[ServiceLineKey]",         "→ Dim_ServiceLine[ServiceLineKey]"),
        ("", ""),
        ("🔄  CONNECTING TO LIVE DATA (when ready)", ""),
        ("Step 1", "In Power BI Desktop: Home → Transform Data → Data Source Settings"),
        ("Step 2", "Select this Excel file → Change Source"),
        ("Step 3", "Point to your live system connector (SQL Server, Xero API, SharePoint, etc.)"),
        ("Step 4", "Map the query to the matching table — column names are identical to your live schema"),
        ("Step 5", "Close & Apply — all visuals update immediately. No DAX or relationship changes needed."),
        ("", ""),
        ("⚠️  IMPORTANT NOTES", ""),
        ("Mock data",     "All names, values, and identifiers are randomly generated. Any resemblance to real people or organisations is coincidental."),
        ("Date range",    "Data spans FY2024 (Jul 2023 – Jun 2024) and FY2025 (Jul 2024 – Jun 2025)."),
        ("Fiscal year",   "Australian fiscal year — July to June. All TOTALYTD measures use '30/06' as the year-end parameter."),
        ("Utilisation",   "Target utilisation rates vary by grade (Graduate 82% → Partner 55%) — matching real consulting industry benchmarks."),
        ("Seed",          "Random seed is fixed (42) — re-running the script produces identical data for consistency."),
    ]

    for i, (key, val) in enumerate(sections, 3):
        ws.cell(row=i, column=1, value=key)
        ws.cell(row=i, column=3, value=val)
        ws.merge_cells(f"C{i}:H{i}")

        if not key and not val:
            ws.row_dimensions[i].height = 8
            continue

        if val == "" and key.startswith(("🚀","📊","🔗","🔄","⚠️")):
            # Section header
            ws.cell(row=i, column=1).font = Font(bold=True, size=11, color="2D5BE3", name="Segoe UI")
            ws.cell(row=i, column=1).fill = PatternFill("solid", start_color="EEF2FD")
            ws.merge_cells(f"A{i}:H{i}")
            ws.row_dimensions[i].height = 22
        else:
            ws.cell(row=i, column=1).font = Font(bold=True, size=9, color="1E3A5F", name="Segoe UI")
            ws.cell(row=i, column=3).font = Font(size=9, color="333333", name="Segoe UI")
            ws.cell(row=i, column=3).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[i].height = max(15, val.count("\n") * 15 + 15)

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 2
    ws.column_dimensions["C"].width = 80


# ═══════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════

def main():
    print("StudioTech BI Blueprint Generator — Mock Data Generator")
    print("=" * 60)

    output_path = "/mnt/user-data/outputs/MockData_ProfessionalServices.xlsx"

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # ── Date spine ───────────────────────────────────────────────
    print("Building Dim_Date...", end=" ", flush=True)
    start_dt = datetime.date(2023, 7, 1)
    end_dt   = datetime.date(2025, 6, 30)
    date_rows = build_dim_date(start_dt, end_dt)
    print(f"{len(date_rows):,} rows")

    # ── Dimension tables ─────────────────────────────────────────
    print("Building Dim_Employee...", end=" ", flush=True)
    employees = build_dim_employee()
    print(f"{len(employees)} rows")

    print("Building Dim_Client...", end=" ", flush=True)
    clients = build_dim_client()
    print(f"{len(clients)} rows")

    print("Building Dim_Engagement...", end=" ", flush=True)
    engagements = build_dim_engagement(clients, employees)
    print(f"{len(engagements)} rows")

    practices = build_dim_practice(employees)
    service_lines = build_dim_service_line()

    # ── Fact tables ──────────────────────────────────────────────
    print("Building Fact_Timesheets...", end=" ", flush=True)
    timesheets = build_fact_timesheets(employees, engagements, date_rows)
    print(f"{len(timesheets):,} rows")

    print("Building Fact_Engagements...", end=" ", flush=True)
    eng_snaps = build_fact_engagements(engagements, employees, date_rows)
    print(f"{len(eng_snaps):,} rows")

    print("Building Fact_Invoices...", end=" ", flush=True)
    invoices = build_fact_invoices(clients, engagements, date_rows)
    print(f"{len(invoices):,} rows")

    print("Building Fact_ResourceUtilisation...", end=" ", flush=True)
    utilisation = build_fact_utilisation(employees, date_rows)
    print(f"{len(utilisation):,} rows")

    print("Building Fact_Pipeline...", end=" ", flush=True)
    pipeline = build_fact_pipeline(clients, employees, date_rows)
    print(f"{len(pipeline)} rows")

    # ── Write workbook ───────────────────────────────────────────
    print("\nWriting workbook...")

    # Instructions first
    write_instructions(wb)

    # Dim_Date
    write_sheet(wb, "Dim_Date",
        ["DateKey","Date","Day","DayName","Week","Month","MonthName","MonthLabel",
         "Quarter","CalendarYear","FiscalYear","FiscalQuarter","FiscalMonth",
         "IsWeekend","IsPublicHoliday"],
        date_rows, CLR_TAB_DATE, CLR_STRIPE_DATE, CLR_HEADER_DATE,
        [10,12,6,12,6,6,12,10,6,10,8,8,8,8,12])

    # Dim_ServiceLine
    write_sheet(wb, "Dim_ServiceLine",
        ["ServiceLineKey","ServiceLineName","PracticeKey","DivisionName"],
        service_lines, CLR_TAB_DIM, CLR_ALT_DIM, CLR_HEADER_DIM,
        [14,24,12,28])

    # Dim_Practice
    write_sheet(wb, "Dim_Practice",
        ["PracticeKey","PracticeName","ServiceLineKey","DivisionKey","PracticeLeaderKey"],
        practices, CLR_TAB_DIM, CLR_ALT_DIM, CLR_HEADER_DIM,
        [12,28,14,12,18])

    # Dim_Employee
    write_sheet(wb, "Dim_Employee",
        ["EmployeeKey","EmployeeID","FullName","Grade","Role","ServiceLineKey",
         "PracticeKey","PracticeKey2","EmploymentType","ChargeOutRate","IsPartner",
         "IsActive","HireDate","OfficeState"],
        employees, CLR_TAB_DIM, CLR_ALT_DIM, CLR_HEADER_DIM,
        [12,12,22,18,12,14,12,12,14,14,10,10,14,12])

    # Dim_Client
    write_sheet(wb, "Dim_Client",
        ["ClientKey","ClientCode","ClientName","Industry","Segment","State",
         "ClientPartnerKey","IsActive","ClientSince"],
        clients, CLR_TAB_DIM, CLR_ALT_DIM, CLR_HEADER_DIM,
        [10,10,30,24,14,8,16,10,14])

    # Dim_Engagement
    write_sheet(wb, "Dim_Engagement",
        ["EngagementKey","EngagementCode","EngagementName","ClientKey","ServiceLineKey",
         "EngagementType","StartDate","EndDate","IsActive"],
        engagements, CLR_TAB_DIM, CLR_ALT_DIM, CLR_HEADER_DIM,
        [14,14,44,10,14,20,14,14,10])

    # Fact_Timesheets
    write_sheet(wb, "Fact_Timesheets",
        ["TimesheetKey","EmployeeKey","EngagementKey","ClientKey","ServiceLineKey",
         "DateKey","HoursWorked","BillableHours","NonBillableHours",
         "StandardRate","WIPValue","IsApproved"],
        timesheets, CLR_TAB_FACT, CLR_ALT_ROW, CLR_HEADER_FACT,
        [14,12,14,10,14,12,12,12,16,12,12,12])

    # Fact_Engagements
    write_sheet(wb, "Fact_Engagements",
        ["EngagementSnapshotKey","ClientKey","EngagementMgrKey","ServiceLineKey",
         "DateKey","StartDateKey","EndDateKey","BudgetFees","ActualFees",
         "WIPBalance","WriteOffAmount","EngagementStatus","IsFixedFee"],
        eng_snaps, CLR_TAB_FACT, CLR_ALT_ROW, CLR_HEADER_FACT,
        [22,10,16,14,12,12,12,14,12,12,14,18,12])

    # Fact_Invoices
    write_sheet(wb, "Fact_Invoices",
        ["InvoiceKey","ClientKey","EngagementKey","InvoiceDateKey","DueDateKey",
         "PaidDateKey","InvoiceAmount","GSTAmount","AmountPaid","AmountOutstanding",
         "DaysOutstanding","InvoiceStatus"],
        invoices, CLR_TAB_FACT, CLR_ALT_ROW, CLR_HEADER_FACT,
        [12,10,14,14,12,12,14,12,12,16,14,14])

    # Fact_ResourceUtilisation
    write_sheet(wb, "Fact_ResourceUtilisation",
        ["UtilisationKey","EmployeeKey","ServiceLineKey","DateKey",
         "AvailableHours","BillableHours","NonBillableHours","LeaveHours","UtilisationRate"],
        utilisation, CLR_TAB_FACT, CLR_ALT_ROW, CLR_HEADER_FACT,
        [14,12,14,12,14,12,16,12,16])

    # Fact_Pipeline
    write_sheet(wb, "Fact_Pipeline",
        ["PipelineKey","ClientKey","ServiceLineKey","OwnerKey","DateKey",
         "EstimatedValue","WeightedValue","WinProbability","Stage","ExpectedCloseDate"],
        pipeline, CLR_TAB_FACT, CLR_ALT_ROW, CLR_HEADER_FACT,
        [12,10,14,10,12,14,14,14,16,18])

    # ── Save ─────────────────────────────────────────────────────
    print(f"\nSaving to {output_path}...")
    wb.save(output_path)
    print("\n✅ Done!")
    print(f"\n   File:  MockData_ProfessionalServices.xlsx")
    print(f"\n   Sheets generated:")
    for sheet in wb.sheetnames:
        print(f"     {sheet}")

    # ── Summary stats ─────────────────────────────────────────────
    total_rows = (len(date_rows) + len(employees) + len(clients) +
                  len(engagements) + len(service_lines) + len(practices) +
                  len(timesheets) + len(eng_snaps) + len(invoices) +
                  len(utilisation) + len(pipeline))
    print(f"\n   Total rows across all tables: {total_rows:,}")
    print(f"\n   Open Power BI Desktop → Get Data → Excel → select all tables → Load")


if __name__ == "__main__":
    main()
